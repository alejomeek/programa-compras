"""Constructores de fixtures **sintéticas** para las pruebas del motor.

Regla del plan §16.7, sin excepciones: ninguna prueba lee los archivos
comerciales de la raíz del repositorio (`SDOSXSUC (7).CSV`, `INVEPTOS.XLS`,
`LISTA DE PRECIOS ...xls`). Todo archivo de prueba se construye en memoria
(`BytesIO`) con EAN, nombres y proveedores inventados.
"""

from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Sequence

import pandas as pd
import pytest

# Permite `import engine...` ejecutando pytest desde la raíz del repositorio.
REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))


# --------------------------------------------------------------------------- #
# EAN sintéticos (inventados; no corresponden a ningún producto real)
# --------------------------------------------------------------------------- #

EAN_CON_CERO_INICIAL = "0012345678905"
EAN_NORMAL = "7701234567890"
EAN_CORTO = "12345"
EAN_OTRO = "7709876543210"


# --------------------------------------------------------------------------- #
# SDOSXSUC (CSV Latin-1, separado por ';')
# --------------------------------------------------------------------------- #


def build_sdos_csv(
    rows: Sequence[dict[str, Any]],
    *,
    columns: Sequence[str] | None = None,
    encoding: str = "latin1",
) -> BytesIO:
    """Arma un CSV tipo SDOSXSUC en memoria.

    ``columns`` permite omitir columnas a propósito (por ejemplo, un archivo
    sin ``us06``) para probar el manejo de columnas faltantes.
    """
    frame = pd.DataFrame(list(rows))
    if columns is not None:
        frame = frame.reindex(columns=list(columns), fill_value="")
    text = frame.to_csv(index=False, sep=";")
    return BytesIO(text.encode(encoding))


def sdos_row(
    *,
    codpro: str = "SKU-001",
    nompro: str = "Rompecabezas Sintético",
    valuni: str = "$ 45.900,00",
    codean: str = EAN_NORMAL,
    codea2: str = ".745SINT",
    **inventory: Any,
) -> dict[str, Any]:
    """Una fila de SDOSXSUC con valores sintéticos por defecto."""
    row: dict[str, Any] = {
        "Codpro": codpro,
        "Nompro": nompro,
        "Valuni": valuni,
        "Codean": codean,
        "Codea2": codea2,
    }
    row.update({key: str(value) for key, value in inventory.items()})
    return row


# --------------------------------------------------------------------------- #
# INVEPTOS (.xls legado BIFF, escrito con xlwt)
# --------------------------------------------------------------------------- #

def _require_xlwt():
    """Devuelve el módulo ``xlwt`` o salta la prueba si no está instalado.

    ``xlwt`` es dependencia **solo de pruebas**: es la única forma de generar
    un ``.xls`` legado real (pandas 2.x ya no escribe ese formato) y así
    ejercitar el mismo camino ``xlrd`` que usa producción.
    """
    return pytest.importorskip(
        "xlwt",
        reason="xlwt es necesario para generar .xls legado sintético",
        exc_type=ImportError,
    )


def build_inveptos_xls(
    header: Sequence[str],
    rows: Iterable[Sequence[Any]],
    *,
    encoding: str = "cp1252",
    sheet_name: str = "INVEPTOS",
) -> BytesIO:
    """Arma un ``.xls`` BIFF en memoria con el encabezado y filas dados."""
    xlwt_mod = _require_xlwt()
    workbook = xlwt_mod.Workbook(encoding=encoding)
    sheet = workbook.add_sheet(sheet_name)
    for col, title in enumerate(header):
        sheet.write(0, col, title)
    for row_index, row in enumerate(rows, start=1):
        for col, value in enumerate(row):
            sheet.write(row_index, col, value)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


#: Encabezado mínimo de INVEPTOS con dos pares TISUC/UNSUC.
INVEPTOS_HEADER_2_SUCURSALES: tuple[str, ...] = (
    "CODPRO",
    "COMODI",
    "DETALL",
    "VALCOS",
    "TISUC1",
    "UNSUC1",
    "SDSUC1",
    "TISUC2",
    "UNSUC2",
    "SDSUC2",
    "UNIVTA",
    "FDESDE",
    "FHASTA",
    "CODEAN",
)


def inveptos_row(
    *,
    codpro: str = "SKU-001",
    comodi: str = ".745SINT",
    detall: str = "Rompecabezas Sintético",
    valcos: str = "20000",
    tisuc1: str = "10000",
    unsuc1: str = "10",
    sdsuc1: str = "3",
    tisuc2: str = "10010",
    unsuc2: str = "5",
    sdsuc2: str = "1",
    univta: str = "15",
    fdesde: str = "01-ene-25",
    fhasta: str = "31-ene-25",
    codean: str = EAN_NORMAL,
) -> tuple[Any, ...]:
    """Una fila alineada a :data:`INVEPTOS_HEADER_2_SUCURSALES`."""
    return (
        codpro,
        comodi,
        detall,
        valcos,
        tisuc1,
        unsuc1,
        sdsuc1,
        tisuc2,
        unsuc2,
        sdsuc2,
        univta,
        fdesde,
        fhasta,
        codean,
    )


# --------------------------------------------------------------------------- #
# Lista de precios de proveedor (.xlsx, encabezado posiblemente desplazado)
# --------------------------------------------------------------------------- #


def build_supplier_xlsx(
    header: Sequence[str],
    rows: Iterable[Sequence[Any]],
    *,
    header_row: int = 0,
    preamble: Sequence[Sequence[Any]] | None = None,
) -> BytesIO:
    """Arma un ``.xlsx`` de lista de precios con el encabezado desplazado.

    ``header_row`` es el índice 0-based de la fila donde va el encabezado; las
    filas anteriores se llenan con ``preamble`` (o quedan vacías), simulando
    los logos y títulos que traen las listas reales.
    """
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    filler = list(preamble or [])
    for index in range(header_row):
        values = filler[index] if index < len(filler) else []
        for col, value in enumerate(values, start=1):
            sheet.cell(row=index + 1, column=col, value=value)
    for col, title in enumerate(header, start=1):
        sheet.cell(row=header_row + 1, column=col, value=title)
    for row_index, row in enumerate(rows, start=header_row + 2):
        for col, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col, value=value)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


# --------------------------------------------------------------------------- #
# DataFrames sintéticos (entrada de `engine.imports`)
# --------------------------------------------------------------------------- #
#
# `engine.imports` recibe el DataFrame que ya devolvieron los lectores, así que
# las pruebas de orquestación no necesitan pasar por un archivo real: se arma
# el DataFrame equivalente (todo texto, sin NaN) directamente. Las pruebas que
# sí quieren ejercitar la cadena completa usan los builders de archivo de
# arriba.


def frame_from_rows(
    header: Sequence[str],
    rows: Iterable[Sequence[Any]],
) -> pd.DataFrame:
    """DataFrame con la misma forma que devuelven los lectores: todo texto."""
    data = [[("" if v is None else str(v)) for v in row] for row in rows]
    return pd.DataFrame(data, columns=list(header))


def inveptos_frame(
    rows: Iterable[Sequence[Any]] | None = None,
    *,
    header: Sequence[str] = INVEPTOS_HEADER_2_SUCURSALES,
) -> pd.DataFrame:
    """DataFrame equivalente a la salida de ``read_inveptos``."""
    return frame_from_rows(header, rows if rows is not None else [inveptos_row()])


SDOS_HEADER_2_UBICACIONES: tuple[str, ...] = (
    "Codpro",
    "Nompro",
    "Valuni",
    "Codean",
    "Codea2",
    "us01",
    "us02",
)


def sdos_frame(
    rows: Iterable[dict[str, Any]] | None = None,
    *,
    columns: Sequence[str] = SDOS_HEADER_2_UBICACIONES,
) -> pd.DataFrame:
    """DataFrame equivalente a la salida de ``read_sdos``."""
    records = list(rows) if rows is not None else [sdos_row(us01="4", us02="6")]
    frame = pd.DataFrame(records).reindex(columns=list(columns), fill_value="")
    return frame.fillna("").astype(str)


SUPPLIER_HEADER: tuple[str, ...] = ("EAN-13", "Nombre", "Costo proveedor")


def supplier_frame(rows: Iterable[Sequence[Any]] | None = None) -> pd.DataFrame:
    """DataFrame equivalente a la salida de ``read_supplier_price_list``."""
    default = [(EAN_NORMAL, "Rompecabezas Sintético", "45.900")]
    return frame_from_rows(SUPPLIER_HEADER, rows if rows is not None else default)


# --------------------------------------------------------------------------- #
# Doble de conexión DB-API 2.0 en memoria
# --------------------------------------------------------------------------- #
#
# `engine.persistence` recibe la conexión inyectada y solo usa la interfaz
# DB-API 2.0, así que las pruebas la sustituyen por este doble. NO se conecta a
# ninguna base real (menos aún al proyecto Supabase del usuario): modela lo
# justo para poder afirmar qué quedó *confirmado* y qué se revirtió.


#: Catálogo `locations` sintético: nombre → id. Los nombres son los del
#: catálogo real del motor porque son la llave del cruce; los ids son ficticios.
SYNTHETIC_LOCATION_IDS: dict[str, str] = {
    "Av. 19": "loc-0001",
    "Bulevar": "loc-0002",
    "Calle 74": "loc-0003",
    "Bvista": "loc-0004",
    "Oviedo": "loc-0005",
    "CEDI": "loc-0006",
    "Feria": "loc-0007",
    "Full MercadoLibre": "loc-0008",
    "Bodega Bqlla": "loc-0009",
}


class FakeDatabaseError(RuntimeError):
    """Fallo simulado de la base (equivale a un error de psycopg)."""


class FakeCursor:
    def __init__(self, connection: "FakeConnection") -> None:
        self._connection = connection
        self._rows: list[tuple[Any, ...]] = []
        self.closed = False

    def execute(self, sql: str, params: Any = ()) -> None:
        self._rows = self._connection._run(sql, params)

    def executemany(self, sql: str, rows: Iterable[Any]) -> None:
        self._connection._run(sql, list(rows), many=True)
        self._rows = []

    def fetchone(self) -> tuple[Any, ...] | None:
        return self._rows[0] if self._rows else None

    def fetchall(self) -> list[tuple[Any, ...]]:
        return list(self._rows)

    def close(self) -> None:
        self.closed = True


class FakeConnection:
    """Conexión DB-API falsa que distingue lo confirmado de lo revertido.

    ``pending`` acumula lo ejecutado desde el último ``commit``/``rollback``;
    ``committed`` solo recibe lo que un ``commit`` confirmó. Así una prueba
    puede afirmar la regla del contrato §9.3 —"ningún dato parcial vigente"—
    mirando `committed`, no la lista total de sentencias.
    """

    def __init__(
        self,
        *,
        job_status: str | None = "pending",
        locations: dict[str, str] | None = None,
        max_version: int | None = None,
        superseded_id: str | None = None,
        fail_on: str | Sequence[str] = (),
    ) -> None:
        self.job_status = job_status
        self.locations = (
            SYNTHETIC_LOCATION_IDS if locations is None else dict(locations)
        )
        self.max_version = max_version
        self.superseded_id = superseded_id
        self.fail_on: tuple[str, ...] = (
            (fail_on,) if isinstance(fail_on, str) else tuple(fail_on)
        )
        self.pending: list[tuple[str, Any]] = []
        self.committed: list[tuple[str, Any]] = []
        self.commits = 0
        self.rollbacks = 0
        self.cursors: list[FakeCursor] = []
        self._inserted: dict[str, int] = {}

    # -- interfaz DB-API ---------------------------------------------------- #

    def cursor(self) -> FakeCursor:
        cursor = FakeCursor(self)
        self.cursors.append(cursor)
        return cursor

    def commit(self) -> None:
        self.commits += 1
        self.committed.extend(self.pending)
        self.pending = []

    def rollback(self) -> None:
        self.rollbacks += 1
        self.pending = []

    # -- ayudas para las pruebas -------------------------------------------- #

    def committed_sql(self) -> list[str]:
        return [sql for sql, _ in self.committed]

    def committed_matching(self, fragment: str) -> list[tuple[str, Any]]:
        return [(sql, params) for sql, params in self.committed if fragment in sql]

    def wrote_to(self, table: str) -> bool:
        """¿Quedó confirmada alguna escritura sobre esa tabla?"""
        return any(
            f"INTO {table} " in sql or sql.startswith(f"UPDATE {table} ")
            for sql in self.committed_sql()
        )

    # -- motor mínimo ------------------------------------------------------- #

    def _run(self, sql: str, params: Any, many: bool = False) -> list[tuple[Any, ...]]:
        statement = " ".join(str(sql).split())
        self.pending.append((statement, params))
        for fragment in self.fail_on:
            if fragment in statement:
                raise FakeDatabaseError(
                    f"fallo simulado de la base al ejecutar: {fragment}"
                )
        if many:
            return []
        return self._result_for(statement)

    def _result_for(self, statement: str) -> list[tuple[Any, ...]]:
        if statement.startswith("SELECT status FROM import_jobs"):
            return [] if self.job_status is None else [(self.job_status,)]
        if statement.startswith("SELECT id, name FROM locations"):
            return [(value, name) for name, value in self.locations.items()]
        if statement.startswith("SELECT version FROM price_lists"):
            return [] if self.max_version is None else [(self.max_version,)]
        if "SET status = 'superseded'" in statement:
            return [] if self.superseded_id is None else [(self.superseded_id,)]
        if statement.startswith("INSERT INTO") and statement.endswith("RETURNING id"):
            table = statement.split()[2]
            self._inserted[table] = self._inserted.get(table, 0) + 1
            return [(f"{table}-{self._inserted[table]}",)]
        return []
